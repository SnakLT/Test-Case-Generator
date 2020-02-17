import openpyxl
import random
import os
import json
import datetime
import string
import traceback
import subprocess
import time
import sqlite3
from time import localtime, strftime
from shutil import copyfile
from random import randint
from datetime import datetime, timedelta

batch_loc = r'C:\Users\konygal\Desktop\Python notebook\Auto_case_Generator\Auto_case_gen_v3\excel_orders'       #Order generation location

with open(r'C:\Users\konygal\Desktop\Python notebook\Auto_case_Generator\f_name_list.txt') as random_f_name:    #List of random First names
    random_fname_list = tuple(random_f_name.read().split('\n'))

with open(r'C:\Users\konygal\Desktop\Python notebook\Auto_case_Generator\l_name_list.txt') as random_l_name:    #List of random Last names
    random_lname_list = tuple(random_l_name.read().split('\n'))    

with open(r'C:\Users\konygal\Desktop\Python notebook\Auto_case_Generator\c_name_list.txt') as random_c_name:    #List of random Company names
    random_company_name_list = tuple(random_c_name.read().split('\n'))

with open(r'C:\Users\konygal\Desktop\Python notebook\Auto_case_Generator\st_name_list.txt') as random_st_name:  #List of random Street names
    random_street_name = tuple(random_st_name.read().split('\n'))    

with open(r'C:\Users\konygal\Desktop\Python notebook\Auto_case_Generator\cty_name_list.txt') as random_cty_name: #List of random City names
    random_city_name = tuple(random_cty_name.read().split('\n'))

country_code_dict = {'CH':'7001', 'FR':'7002', 'DE':'7004', 'ES':'7006', #Excel template upload location country code
                     'FI':'7003', 'HU':'7005', 'NO':'7008', 'SE':'7007',
                     'CZ':'7009', 'LV':'7012', 'IE':'7011', 'PT':'7013',
                     'BE':'7019', 'AT':'7017', 'NL':'7020', 'IT':'7021'
                     }

code_len = '0000000000' #Client code length, to generate ordered clients code

def get_info_from_db(json_content):
    
    conn = sqlite3.connect('users.db')
    curr = conn.cursor()
    
    curr.execute('SELECT * FROM total_orders WHERE username=?', (json_content[0]['username'],))
    if len(curr.fetchall()) == 0: #If the user is new, add him to database
        curr.execute('INSERT INTO total_orders (username,order_nr,case_nr) VALUES(?,?,?);',(json_content[0]['username'], 0, 0))
        
    curr.execute('SELECT order_nr,case_nr FROM total_orders WHERE username=?', (json_content[0]['username'],))
    info = curr.fetchall()[0]
    order_nr, line_nr = info
    conn.commit()
    conn.close()
    return order_nr + 1, line_nr

def update_db(json_content, case_nr, order_nr):
    
    conn = sqlite3.connect('users.db')
    curr = conn.cursor()
    curr.execute('UPDATE total_orders SET order_nr=?, case_nr=? WHERE username=?', (order_nr, case_nr, json_content[0]['username'],))
    conn.commit()
    conn.close()    
    

def randomStringDigits(stringLength=12): #Generating random string for invoice number
    lettersAndDigits = string.ascii_letters + string.digits
    return ''.join(random.choice(lettersAndDigits) for i in range(stringLength))


def generate_sum(json_content, submit): #Generates random integer from given range
    return str(random.randint(json_content[submit]['debt_amount'][0], json_content[submit]['debt_amount'][1]))


def read_json_files(file): #Reads users input json order as dictionary

    try:
        with open(file) as json_file:
            json_content = json.load(json_file)
        
        return json_content
    
    except:
        traceback.print_exc()
        

def generate_excel_file(json_content, num, line_nr):

    today = datetime.today()
    order_dict = {} # Dictionary for order monitoring

    for submit in range(len(json_content)):
        
        new_file_name = str(num) + '_' + today.strftime("%m-%d") + '_' + json_content[submit]['username'] \
        + '_' + json_content[submit]['client_nr'] + '_' + json_content[submit]['country'] \
        + '_' + json_content[submit]['instance'] + '.xlsx'

        template_file = f'{batch_loc}\\{new_file_name}'
        
        if new_file_name not in os.listdir(batch_loc):
            copyfile('Template.xlsx', template_file)
        
        wb = openpyxl.load_workbook(filename=template_file)
        ws = wb['1. Template']

        full_acc_ref_list_email = [] #List for building email text
        full_acc_ref_list = []  #List for building case numbers to search for during monitoring

        #Constant variables for one order
        count_inst = f'{json_content[submit]["country"]}-{json_content[submit]["instance"]}'
        nr_of_cases = json_content[submit]['nr_of_cases']
        cust_type = json_content[submit]['customer_type']
        client_nr = country_code_dict[json_content[submit]['country']] + code_len[len(json_content[submit]["client_nr"]):] + json_content[submit]["client_nr"]
        #First line of order text for email
        full_acc_ref_list_email.append(f'{json_content[submit]["country"]} {json_content[submit]["instance"]} Order nr: {submit + 1}; Client number: {client_nr}, {nr_of_cases} {cust_type} cases')
        
        for i in range(nr_of_cases):

            #Variables for each ordered case
            customer_type = '1' if json_content[submit]['customer_type'] == 'B2B' else '2'
            debt_maturity = json_content[submit]['debt_maturity'] if json_content[submit]['debt_maturity'] is not None else 60
            inv_due_date = today - timedelta(days=debt_maturity)
            inv_date = today - timedelta(days=debt_maturity + 30)
            company_name = None if customer_type == '2' else random.choice(random_company_name_list)
            first_name = random.choice(random_fname_list)
            last_name = random.choice(random_lname_list)
            debtor_minor = today - timedelta(days = randint(20, 60) * 365) if json_content[submit]['debtor_minor'] is None else today - timedelta(days = 14 * 365)
            service_type = ('Reminder Service','Debt Collection','Short Amicable') if json_content[submit]['service'] is not None else None
            service_type = service_type.index(json_content[submit]['service']) + 1 if service_type is not None else None
            acc_ref_nr = json_content[submit]['username'] + '-' + json_content[submit]['country'] + '-'
            if i == 0:
                full_acc_ref_list.append(acc_ref_nr + str(line_nr + 1))
            if json_content[submit]['international'] is True:
                fixed_country_dict = country_code_dict.copy()
                if json_content[submit]['country'] in fixed_country_dict:
                    del fixed_country_dict[json_content[submit]['country']]
                country_keys = [key for key in fixed_country_dict.keys()]
                inter_country = random.choice(country_keys)
            else:
                inter_country = json_content[submit]['country']

            new_row = [acc_ref_nr + str(line_nr + 1), 'Invoice', customer_type, '', '', '', company_name, last_name, first_name, '', '', debtor_minor.strftime("%d-%m-%Y"),
                       random.choice(random_street_name) + ' ' + str(randint(1, 99)), '', str(randint(1000, 9999)), random.choice(random_city_name),
                       random.choice(random_city_name) + ' Province', inter_country, '', '', '', '', '', '', '', '', '+' + str(randint(100000000, 999999999)), first_name + '.' + last_name + '@testcase.com',
                       randomStringDigits(), inv_date.strftime("%d-%m-%Y"), inv_due_date.strftime("%d-%m-%Y"), '', generate_sum(json_content, submit),
                       generate_sum(json_content, submit), json_content[submit]['intrest_rate'], json_content[submit]['client_costs'], service_type]
        
            if i == 0: #If its a first case from order, adds text with from case
                full_acc_ref_list_email.append('From: ' + new_row[0])
                

            if json_content[submit]['multiple_cases'] is not None: #if order must contain multiple cases on the same customer
                for _ in range(json_content[submit]['nr_of_cases']):
                    line_nr += 1
                    new_row[0] = acc_ref_nr + str(line_nr)
                    new_row[28] = randomStringDigits()
                    new_row[32] = generate_sum(json_content, submit)
                    new_row[33] = new_row[32]                    
                    ws.append(new_row)
                
                    if json_content[submit]['multiple_debt_items'] is not None: #if order must contain multiple cases on the same customer with multiple debt items
                        
                        for _ in range(json_content[submit]['multiple_debt_items']):
                            new_row[28] = randomStringDigits()
                            new_row[32] = generate_sum(json_content, submit)
                            new_row[33] = new_row[32]
                            ws.append(new_row)
                
                full_acc_ref_list_email.append('To:      ' + new_row[0] + '_ _') #Last line To for email text
                wb.save(template_file)
                break
                
            elif json_content[submit]['multiple_debt_items'] is not None: #if order must contain multiple debt items
                line_nr += 1
                for _ in range(json_content[submit]['multiple_debt_items']):
                    new_row[28] = randomStringDigits()
                    new_row[32] = generate_sum(json_content, submit)
                    new_row[33] = new_row[32]
                    ws.append(new_row)
                
            else:
                line_nr += 1 
                ws.append(new_row)
                
                
            if i == json_content[submit]['nr_of_cases'] - 1: #Last line To for email text
                full_acc_ref_list_email.append(f'To:      {new_row[0]}_ _')


        name = json_content[submit]['country'] + '_' + json_content[submit]['instance']
        if name not in order_dict: #Building a dictionary for order monitoring
            order_dict[name] = {}
            order_dict[name]['accref_email'] = [full_acc_ref_list_email]
            order_dict[name]['accref'] = full_acc_ref_list
            order_dict[name]['nr_cases'] = [nr_of_cases]
            
        else:
            order_dict[name]['accref'].extend(full_acc_ref_list)
            order_dict[name]['accref_email'].append(full_acc_ref_list_email)
            order_dict[name]['nr_cases'].append(nr_of_cases)
    
        wb.save(template_file)
    
    return line_nr, order_dict


def upload_excel_file(check_dict):

    nr_of_files = [] #File names for order monitoring

    for file in os.listdir(batch_loc):
        file_name = file[:-5].split('_')
           
        if file_name[-1] == 'TEST':        
            upload_link = '\\\\FIHEL2STAS145.groupad1.com\\Scale\\New-Case\\'
        else:
            upload_link = '\\\\DEFRA1SUAS024.groupad1.com\\Scale\\New-Case\\'
        
        country = file_name[-2]
        client_nr = country_code_dict[country] + code_len[len(file_name[-3]):] + file_name[-3]
        final_link = f'{upload_link}{country}\\{client_nr}\\in\\{file}'
        file_loc = f'{batch_loc}\\{file}'
        try:
            copyfile(file_loc, final_link)
            os.remove(file_loc)
            nr_of_files.append(file)
        except: #If client is not present, checkes order number and builds email text
            print('No such Client')
            wb = openpyxl.load_workbook(filename=batch_loc + '\\' + file)
            ws = wb['1. Template']
            acc_ref = [cell.value for cell in ws['A']][2]
            for order in check_dict:
                for ref in range(len(check_dict[order]['accref'])):
                    if acc_ref in check_dict[order]['accref'][ref]:
                        check_dict[order]['accref'][ref] = 'PASS'
                        check_dict[order]['accref_email'][ref] = [check_dict[order]['accref_email'][ref][0], 'Client is not onboarded in CAR _ _']
            os.remove(f'{batch_loc}\\{file}')
            continue
        
    return nr_of_files, check_dict

    
def start_work(file):

    try:
        json_content = read_json_files(file) #Receives file information from Main execution
        order_nr, case_nr = get_info_from_db(json_content) #Receives latest order number and last used case number
        case_nr, check_dict = generate_excel_file(json_content, order_nr, case_nr) #Generates Excel files with all orders in {batch_loc} folder and builds a dictionary required for order monitoring, text for email and final case number. 
        nr_of_files, check_dict = upload_excel_file(check_dict) #Uploads all Excel files from {batch_loc} and updates monitoring dictionary if client number is not awailable
        update_db(json_content, case_nr, order_nr) #Updates the database with current order number and final case number
        json_transfer = {'user': [json_content[0]["username"], json_content[0]["email"], nr_of_files], #Builds final dictionary for order monitoring
                         'order': check_dict}
        
        return json_transfer

    except Exception:
        traceback.print_exc()



